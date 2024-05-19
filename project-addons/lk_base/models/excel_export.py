# Copyright 2019 Ecosoft Co., Ltd (http://ecosoft.co.th/)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html)

import base64
import logging
import os
import time
from datetime import date, datetime as dt
from io import BytesIO
import json

from odoo import _, api, fields, models
from odoo.exceptions import ValidationError
from odoo.addons.lk_base.models import general_function as gf

from copy import copy

from . import excel_common as co

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
    from openpyxl.utils.cell import column_index_from_string
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side, Color, GradientFill, Fill, Protection
except ImportError:
    _logger.debug('Cannot import "openpyxl". Please make sure it is installed.')


class ExcelExport(models.AbstractModel):
    _name = "excel.export"

    @api.model
    def apply_style_merged_cell(self, ws):
        # fix error of openpyxl to set style of merge cell
        for merged_cells in ws.merged_cells.ranges:
            style = ws.cell(merged_cells.min_row, merged_cells.min_col)._style
            for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                    ws.cell(row, col)._style = style

    @api.model
    def apply_cell_style(self, st, data):
        if not (data.get('excel_cell')):
            raise ValidationError('Missing required attribute: excel_cell')
        cell = st[data['excel_cell']]
        if data.get('copy_format'):
            col, row = co.split_row_col(data['copy_format'])
            copy_cell = st.cell(row=row, column=column_index_from_string(col))
            if copy_cell.has_style:
                cell.font = copy(copy_cell.font)
                cell.border = copy(copy_cell.border)
                cell.fill = copy(copy_cell.fill)
                cell.number_format = copy(copy_cell.number_format)
                cell.protection = copy(copy_cell.protection)
                cell.alignment = copy(copy_cell.alignment)
        if data.get('font'):
            font = copy(cell.font)
            if not isinstance(data['font'], dict):
                raise ValidationError('Format of font: "font": {"bold": True}')
            for key, value in data['font'].items():
                excel_key = key
                if key == 'size':
                    excel_key = 'sz'
                if key == 'bold':
                    excel_key = 'b'
                if key == 'italic':
                    excel_key = 'i'
                if key == 'strikethrough':
                    excel_key = 'strike'
                if key == 'underline':
                    excel_key = 'u'
                setattr(font, excel_key, value)
            cell.font = font
        if data.get('border'):
            border = copy(cell.border)
            if not isinstance(data['border'], dict):
                raise ValidationError('Format of border: "border": {"left": {"style": "thin"}}')
            for key, value in data['border'].items():
                side = Side()
                if not isinstance(value, dict):
                    raise ValidationError('Format of border: "border": {"left": {"style": "thin"}}')
                for child_key, child_value in value.items():
                    excel_key = child_key
                    if key == 'border_style':
                        excel_key = 'style'
                    setattr(side, excel_key, child_value)
                setattr(border, key, side)
            cell.border = border
        if data.get('pattern_fill'):
            fill = copy(cell.fill)
            if not isinstance(data['pattern_fill'], dict):
                raise ValidationError('Format of pattern fill: "pattern_fill": {"start_color": "FFFF0000"}')
            for key, value in data['pattern_fill'].items():
                excel_key = key
                if key == 'fill_type':
                    excel_key = 'patternType'
                if key == 'start_color':
                    excel_key = 'fgColor'
                if key == 'end_color':
                    excel_key = 'bgColor'
                setattr(fill, excel_key, value)
            cell.fill = fill
        if data.get('number_format'):
            if not isinstance(data['number_format'], str):
                raise ValidationError('Format of number_format: "number_format": "0.00"')
            cell.number_format = data['number_format']
        if data.get('protection'):
            protection = copy(cell.protection)
            if not isinstance(data['protection'], dict):
                raise ValidationError('Format of protection: "protection": {"locked": True, "hidden": False}')
            for key, value in data['protection'].items():
                setattr(protection, key, value)
            cell.protection = protection
        if data.get('alignment'):
            alignment = copy(cell.alignment)
            if not isinstance(data['alignment'], dict):
                raise ValidationError('Format of alignment: "alignment": {"horizontal": "center"}')
            for key, value in data['alignment'].items():
                excel_key = key
                if key == 'text_rotation':
                    excel_key = 'textRotation'
                if key == 'wrap_text':
                    excel_key = 'wrapText'
                if key == 'shrink_to_fit':
                    excel_key = 'shrinkToFit'
                setattr(alignment, excel_key, value)
            cell.alignment = alignment

    @api.model
    def _get_field_data(self, _field, _line):
        """ Get field data, and convert data type if needed """
        if not _field:
            return None
        line_copy = _line
        for f in _field.split("."):
            field_attrs = dict(self.env[line_copy._name].fields_get([f]))[f]
            if field_attrs['type'] == 'selection' and line_copy[f]:
                line_copy = dict(self.env[line_copy._name]._fields[f]._description_selection(self.env))[line_copy[f]]
            elif field_attrs['type'] == 'datetime':
                line_copy = gf.get_valid_data(line_copy[f], datetime_format='%d-%m-%Y %H:%M:S')
            elif field_attrs['type'] == 'date':
                line_copy = gf.get_valid_data(line_copy[f], datetime_format='%d-%m-%Y')
            else:
                line_copy = line_copy[f]
        line_copy = gf.get_valid_data(line_copy)
        if isinstance(line_copy, str):
            line_copy = line_copy.encode("utf-8")
        return line_copy

    @api.model
    def _get_line_vals(self, record, line_field, fields):
        """ Get values of this field from record set and return as dict of vals
            - record: main object
            - line_field: rows object, i.e., line_ids
            - fields: fields in line_ids, i.e., partner_id.display_name
        """
        line_field, max_row = co.get_line_max(line_field)
        lines = record[line_field]
        if len(lines) > max_row > 0:
            raise Exception(_("Records in %s exceed max records allowed") % line_field)
        vals = {field: [] for field in fields}
        aggre_func_dict = {}
        merge_dict = {}
        pair_fields = []  # I.e., ('debit${value and . or .}@{sum}', 'debit')
        for field in fields:
            temp_field, field_merge = co.get_field_merge(field)
            raw_field, aggre_func = co.get_field_aggregation(temp_field)
            merge_dict.update({field: field_merge})
            aggre_func_dict.update({field: aggre_func})
            pair_fields.append((field, raw_field))
        for line in lines:
            for field in pair_fields:
                merge_val = merge_dict[field[0]]
                value = self._get_field_data(field[1], line)
                vals[field[0]].append((value, merge_val))
        return (vals, aggre_func_dict)

    @api.model
    def _fill_lines(self, ws, st, record, manual_data=False):
        row_to_copy = 0
        copy_lst = []
        insert_row = []
        if manual_data:
            for record in json.loads(manual_data):
                if record.get('row'):
                    # I forgot why to do this, I think it is function about duplicate row
                    # last_row = record.get('row')[-1]
                    # last_col, last_row = co.split_row_col(last_row['excel_cell'])
                    for data in record.get('row'):
                        if not (data.get('excel_cell') or data.get('value')):
                            raise ValidationError('Missing required attribute: excel_cell / value')
                        self.apply_cell_style(st, data)
                        # col, row = co.split_row_col(data['excel_cell'])
                        # insert new excel row equal to data length
                        # if row not in insert_row and row != last_row:
                        #     st.insert_rows(row + 1)
                        #     insert_row.append(row)
                        # if not row_to_copy:
                        #     row_to_copy = row
                        if data.get('merge_cell'):
                            st.merge_cells('%s:%s' % (data['excel_cell'], data['merge_cell']))
                        st[data['excel_cell']] = co.str_to_number(data['value'])
                        # new_cell = st.cell(row=row, column=column_index_from_string(col))
                        # # copy format previous row
                        # if row > row_to_copy:
                        #     copy_cell = st.cell(row=row_to_copy, column=column_index_from_string(col))
                        #     if copy_cell.has_style:
                        #         new_cell.font = copy(copy_cell.font)
                        #         new_cell.border = copy(copy_cell.border)
                        #         new_cell.fill = copy(copy_cell.fill)
                        #         new_cell.number_format = copy(copy_cell.number_format)
                        #         new_cell.protection = copy(copy_cell.protection)
                        #         new_cell.alignment = copy(copy_cell.alignment)
                if record.get('header'):
                    for data in record.get('header'):
                        if not (data.get('excel_cell') or data.get('value')):
                            raise ValidationError('Missing required attribute: excel_cell / value')
                        self.apply_cell_style(st, data)
                        if data.get('merge_cell'):
                            st.merge_cells('%s:%s' % (data['excel_cell'], data['merge_cell']))
                        st[data['excel_cell']] = data['value']
        else:
            line_fields = list(ws)
            if "_HEAD_" in line_fields:
                line_fields.remove("_HEAD_")
            for line_field in line_fields:
                fields = ws.get(line_field, {}).values()
                vals, func = self._get_line_vals(record, line_field, fields)
                for rc, field in ws.get(line_field, {}).items():
                    col, row = co.split_row_col(rc)  # starting point
                    if not row_to_copy:
                        row_to_copy = row
                    i = 0
                    new_row = 0
                    new_rc = False
                    for (row_val, merge_val) in vals[field]:
                        new_row = row + i
                        new_rc = "{}{}".format(col, new_row)
                        # insert new excel row equal to data length
                        if new_row not in insert_row:
                            if i != len(vals[field]) - 1:
                                st.insert_rows(new_row + 1)
                                insert_row.append(new_row)
                        if merge_val:
                            merge_col, merge_row = co.split_row_col(merge_val)
                            if new_row > row_to_copy:
                                # merge col E6:F6, next cell E7:F7
                                if merge_row == new_row:
                                    new_merge = new_row
                                else:
                                    # merge col and row E6:F7, next cell E8:F9
                                    # merge row E6:E7, next cell E8:E9
                                    new_merge = new_row + new_row - merge_row
                                new_merge_row = "{}{}".format(merge_col, new_merge)
                                st.merge_cells('%s:%s' % (new_rc, new_merge_row))
                            else:
                                st.merge_cells('%s:%s' % (new_rc, merge_val))
                        row_val = co.adjust_cell_formula(row_val, i)
                        if row_val not in ("None", None):
                            st[new_rc] = co.str_to_number(row_val)
                        # copy format previous row
                        if new_row > row_to_copy and new_row not in copy_lst:
                            for index, val in enumerate(st.iter_rows(row_to_copy, row_to_copy)):
                                for cell in val:
                                    new_cell = st.cell(row=new_row, column=cell.column)
                                    if cell.has_style:
                                        new_cell.font = copy(cell.font)
                                        new_cell.border = copy(cell.border)
                                        new_cell.fill = copy(cell.fill)
                                        new_cell.number_format = copy(cell.number_format)
                                        new_cell.protection = copy(cell.protection)
                                        new_cell.alignment = copy(cell.alignment)
                            copy_lst.append(new_row)
                        if merge_val:
                            merge_col, merge_row = co.split_row_col(merge_val)
                            if merge_row == new_row:
                                i += 1
                            else:
                                i += merge_row - new_row + 1
                        else:
                            i += 1
                    # Add footer line if at least one field have sum
                    f = func.get(field, False)
                    if f and new_row > 0:
                        new_row += 1
                        f_rc = "{}{}".format(col, new_row)
                        st[f_rc] = "={}({}:{})".format(f, rc, new_rc)
                        styles = self.env["excel.style"].get_openpyxl_styles()
                        co.fill_cell_style(st[f_rc], "#{font=bold;fill=grey;align=center;style=number}", styles)
            return

    @api.model
    def _fill_head(self, ws, st, records, manual_data=False):
        if not manual_data:
            for rc, field in ws.get("_HEAD_", {}).items():
                temp_field, field_merge = co.get_field_merge(field)
                value = temp_field and self._get_field_data(temp_field, records)
                if field_merge:
                    st.merge_cells('%s:%s' % (rc, field_merge))
                if value is not None:
                    st[rc] = value
        else:
            if manual_data:
                for record in json.loads(manual_data):
                    # set width for column
                    if record.get('col_width'):
                        for data in record['col_width']:
                            for col, col_width in data.items():
                                st.column_dimensions[col].width = col_width

    @api.model
    def _fill_workbook_data(self, workbook, records, data_dict, template):
        """ Fill data from record with style in data_dict to workbook """
        if not template.manual_data and not data_dict:
            return
        try:
            count = 1
            if len(workbook.worksheets) != 1:
                raise Exception(_("Excel need contain one worksheet"))
            worksheet_template = workbook.worksheets[0]
            worksheet_template.title = 'Template'
            if template.per_sheet:
                for record in records:
                    st = workbook.copy_worksheet(worksheet_template)
                    st.title = 'Sheet%s' % str(count)
                    # Fill data, header and rows
                    self._fill_head(data_dict, st, record, manual_data=template.manual_data)
                    self._fill_lines(data_dict, st, record, manual_data=template.manual_data)
                    self.apply_style_merged_cell(st)
                    count += 1
            else:
                st = workbook.copy_worksheet(worksheet_template)
                st.title = 'Sheet1'
                # Fill data, header and rows
                self._fill_head(data_dict, st, records, manual_data=template.manual_data)
                self._fill_lines(data_dict, st, records, manual_data=template.manual_data)
                self.apply_style_merged_cell(st)
        except KeyError as e:
            raise ValidationError(_("Key Error\n%s") % e)
        except IllegalCharacterError as e:
            raise ValidationError(_("IllegalCharacterError\n Some exporting data contain special character\n%s") % e)
        except Exception as e:
            raise ValidationError(_("Error filling data into Excel sheets\n%s") % e)

    @api.model
    def export_xlsx(self, template_name, res_model, res_ids):
        template = self.env['excel.template'].search([('model_id.model', '=', res_model),
                                                      ('name', '=', template_name)], limit=1)
        if not template:
            raise ValidationError(_("Excel template for model %s not found") % res_model)
        data_dict = co.literal_eval(template.instruction.strip())
        export_dict = data_dict.get("__EXPORT__", False)
        out_name = template.report_name
        if not export_dict and not template.manual_data:  # If there is not __EXPORT__ formula, just export
            out_name = template.fname
            out_file = template.datas
            return (out_file, out_name)
        # Prepare temp file (from now, only xlsx file works for openpyxl)
        decoded_data = base64.decodebytes(template.datas)
        stamp = dt.utcnow().strftime("%H%M%S%f")[:-3]
        ftemp = "temp{}.xlsx".format(stamp)
        f = open(ftemp, "wb")
        f.write(decoded_data)
        f.seek(0)
        f.close()
        # Workbook created, temp file removed
        wb = load_workbook(ftemp)
        os.remove(ftemp)
        # Start working with workbook
        records = res_model and self.env[res_model].browse(res_ids) or False
        self._fill_workbook_data(wb, records, export_dict, template)
        # remove worksheet template
        wb.remove_sheet(wb.worksheets[0])
        # Return file as .xlsx
        content = BytesIO()
        wb.save(content)
        content.seek(0)  # Set index to 0, and start reading
        out_file = base64.encodebytes(content.read())
        fname = out_name.replace(" ", "").replace("/", "")
        ts = fields.Datetime.context_timestamp(self, dt.now())
        out_name = "{}_{}".format(fname, ts.strftime("%Y%m%d"))
        if not out_name or len(out_name) == 0:
            out_name = "noname"
        out_ext = "xlsx"
        return (out_file, "{}.{}".format(out_name, out_ext))
